import os
import json
from concurrent.futures import ThreadPoolExecutor
import time
from mcp.server.fastmcp import FastMCP
import re
import PyPDF2
from docx import Document
import openpyxl
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity


class FileSearcherByContent:
    def __init__(self):
        self.max_fsize = 100 * 1024 * 1024
        self.max_prev = 500
        try:
            self.semantic_model = SentenceTransformer('all-MiniLM-L6-v2')
        except:
            self.semantic_model = None
    
    def extracttext_fromfile(self, file_path: str, max_chars: int = 2000) -> str:
        """Extract text from various file types"""
        plaintext_extensions = [
            '.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml',
            '.yaml', '.yml', '.ini', '.log', '.csv', '.ts', '.jsx', '.tsx',
            '.sh', '.bat', '.java', '.c', '.cpp', '.h', '.hpp', '.sql', '.php'
        ]
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == '.pdf':
                return self.extract_pdf(file_path, max_chars)
            elif file_ext in ['.docx', '.doc']:
                return self.extract_docx(file_path, max_chars)
            elif file_ext in ['.xlsx', '.xls']:
                return self.extract_excel(file_path, max_chars)
            elif file_ext in plaintext_extensions:
                return self.extract_plain(file_path, max_chars)
            else:
                return ""
        except Exception as e:
            return ""
    
    def extract_pdf(self, file_path: str, max_chars: int) -> str:
        """Extract text from PDF files"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page_num in range(min(5, len(pdf_reader.pages))):
                    text += pdf_reader.pages[page_num].extract_text()
                    if len(text) > max_chars:
                        break
                return text[:max_chars]
        except:
            return ""
    
    def extract_docx(self, file_path: str, max_chars: int) -> str:
        """Extract text from Word documents"""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
                if len(text) > max_chars:
                    break
            return text[:max_chars]
        except:
            return ""
    
    def extract_excel(self, file_path: str, max_chars: int) -> str:
        """Extract text from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            text = ""
            for sheet_name in list(workbook.sheetnames)[:3]:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(max_row=50, values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
                    if len(text) > max_chars:
                        break
            return text[:max_chars]
        except:
            return ""
    
    def extract_plain(self, file_path: str, max_chars: int) -> str:
        """Extract text from plain text files"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                return file.read(max_chars)
        except:
            try:
                with open(file_path, 'r', encoding='latin-1', errors='ignore') as file:
                    return file.read(max_chars)
            except:
                return ""
    
    def make_keywords(self, content_hint: str) -> list[str]:
        """Extract meaningful keywords from content hint"""
        stopper = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'about'}
        words = re.findall(r'\b\w+\b', content_hint.lower())
        keywords = [word for word in words if len(word) > 2 and word not in stopper]
        return keywords
    
    def calculate_keyword_score(self, keywords: list[str], text: str) -> float:
        """Calculate keyword matching score"""
        if not keywords or not text:
            return 0.0
        text__ = text.lower()
        matches = sum(1 for keyword in keywords if keyword in text__)
        return matches / len(keywords)
    
    def calculate_semantic_score(self, content_hint: str, text: str) -> float:
        """Calculate semantic similarity score"""
        if not self.semantic_model or not text.strip():
            return 0.0
        
        try:
            hint_embedding = self.semantic_model.encode([content_hint])
            text_embedding = self.semantic_model.encode([text[:1000]])
  
            similarity = cosine_similarity(hint_embedding, text_embedding)[0][0]
            return float(similarity)
        except Exception:
            return 0.0
    
    def calculate_relevance_score(self, content_hint: str, extracted_text: str, filename: str) -> float:
        """Calculate overall relevance score"""
        keywords = self.make_keywords(content_hint)
        
        keyword_score = self.calculate_keyword_score(keywords, extracted_text)

        filename_score = self.calculate_keyword_score(keywords, filename)
        if self.semantic_model:
            semantic_score = self.calculate_semantic_score(content_hint, extracted_text)
            return (keyword_score * 0.35) + (semantic_score * 0.5) + (filename_score * 0.15)
        else:
            return (keyword_score * 0.8) + (filename_score * 0.2)
    
    def eligible_files(self, drive: str, extension: str) -> list[str]:
        """Get list of candidate files efficiently"""
        eligible = []
        drive_path = f"{drive}:" if len(drive) == 1 else drive
        
        def scan_directory(directory: str, depth: int = 0):
            """Recursively scan directory with depth limit"""
            if depth > 6:
                return
            try:
                with os.scandir(directory) as entries:
                    for entry in entries:
                        if entry.is_file():
                            if entry.name.lower().endswith(f".{extension.lower()}"):
                                if entry.stat().st_size < self.max_fsize:
                                    eligible.append(entry.path)
                        elif entry.is_dir() and not entry.name.startswith('.'):
                            scan_directory(entry.path, depth + 1)
            except (PermissionError, OSError):
                pass 

        if os.path.exists(drive_path):
            scan_directory(drive_path)
        
        return eligible
    
    async def search_files(self, drive: str, extension: str, content_hint: str, max_results: int = 10):
        """Main search function"""
        eligible = self.eligible_files(drive, extension)
        if not eligible:
            return []
        
        results = []
        processed_count = 0
        error_count = 0
        with ThreadPoolExecutor(max_workers=4) as executor:
            batch_size = 20
    
            for num, i in enumerate(range(0, len(eligible), batch_size), 1):
                batch = eligible[i:i + batch_size]
                futures = []
                for file_path in batch:
                    future = executor.submit(self._process_single_file, file_path, content_hint)
                    futures.append((future, file_path))
    
                batch_results = 0
                for future, file_path in futures:
                    try:
                        result = future.result(timeout=20)
                        processed_count += 1
                        if result is not None:
                            if result['relevance_score'] > 0.3:
                                results.append(result)
                                batch_results += 1
                        
                    except TimeoutError:
                        error_count += 1
                    except Exception as e:
                        error_count += 1
                
        results.sort(key=lambda x: x['relevance_score'], reverse=True)
        return results[:max_results]

    def _process_single_file(self, file_path: str, content_hint: str) -> dict | None:
        """Process a single file for relevance"""
        try:
            stat = os.stat(file_path)
            filename = os.path.basename(file_path)

            extracted_text = self.extracttext_fromfile(file_path)
            
            if not extracted_text:
                keywords = self.make_keywords(content_hint)
                relevance_score = self.calculate_keyword_score(keywords, filename) * 0.5
            else:
                relevance_score = self.calculate_relevance_score(content_hint, extracted_text, filename)
                
            preview = extracted_text[:self.max_prev] if extracted_text else f"File: {filename}"
            
            return {
                    "path": file_path,
                    "filename": filename,
                    "size": stat.st_size,
                    "modified": stat.st_mtime,
                    "relevance_score": relevance_score,
                    "preview": preview}
        
        except Exception:
            return None

# MCP
mcp = FastMCP("FindFileByContent", dependencies=["sentence_transformers", "PyPDF2", "python-docx", "openpyxl"])
searcher = FileSearcherByContent()

@mcp.tool(description="Search for files on specified drive with given extension that match the content hint and display filename and its path of the file having greater relevance score. For example: 'Search for files on E drive with pdf extension that is about Breadth first search' and display filename and path of the file having greater relevance score.")
async def search_files_by_content(drive: str, extension: str, content_hint: str) -> str:
    """
    Search for files on specified drive with given extension that match the content hint.
    Args:
        drive: Drive letter (e.g., 'C', 'D', 'E') or full path
        extension: File extension to search for (e.g., 'pdf', 'docx', 'txt')
        content_hint: Description of the content to search for
    Returns:
        JSON string with search results having filename, path, size, modified, relevance score and preview
    """
    try:
        results = await searcher.search_files(drive, extension, content_hint)
        json_results = []
        for result in results:
            if result['relevance_score'] > 0.4:
                json_results.append({
                "path": result['path'],
                "filename": result['filename'],
                "size": result['size'],
                "modified": time.ctime(result['modified']),
                "relevance_score": round(result['relevance_score'], 3),
                "preview": result['preview']
                })
                
        """
        Display the filename and path of the matching files having greater relevance score
        """
        return json.dumps({
            "status": "success",
            "found": len(json_results),
            "results": json_results
        }, indent=2)
    
    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e),
            "results": []
        }, indent=2)


@mcp.tool(description="List all available disk drives on the system.")
def list_drives() -> str:
    """Return a list of available disk drives."""
    drives = [f"{chr(d)}:" for d in range(65, 91) if os.path.exists(f"{chr(d)}:\\")]
    return "\n".join(drives) if drives else "No drives available."
